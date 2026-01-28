import openpyxl
import os
import json

def analyze_excel_file(file_path):
    """Excel dosyasını analiz eder ve formülleri çıkarır"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {os.path.basename(file_path)}")
    print('='*80)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)

        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]

            # Her hücreyi tara
            formulas = []
            data = []

            for row_idx, row in enumerate(ws.iter_rows(), 1):
                row_data = []
                for col_idx, cell in enumerate(row, 1):
                    if cell.value:
                        cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"

                        # Formül mü kontrol et
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append({
                                'cell': cell_ref,
                                'formula': cell.value,
                                'row': row_idx,
                                'col': col_idx
                            })
                            row_data.append(f"{cell_ref}: {cell.value}")
                        else:
                            row_data.append(f"{cell_ref}: {cell.value}")

                if row_data:
                    data.append(row_data)

            # İlk 20 satırı göster
            print("\nFirst 20 rows of data:")
            for i, row in enumerate(data[:20], 1):
                print(f"Row {i}: {row}")

            # Formülleri göster
            if formulas:
                print(f"\n{len(formulas)} formulas found:")
                for f in formulas[:30]:  # İlk 30 formül
                    print(f"  {f['cell']}: {f['formula']}")
            else:
                print("\nNo formulas found in this sheet.")

        wb.close()

    except Exception as e:
        print(f"Error analyzing {file_path}: {str(e)}")

# Excel dosyalarını analiz et
excel_files = [
    "/home/user/BorsaPortal/Temel/DEĞER+BİÇME.xlsx",
    "/home/user/BorsaPortal/Temel/1.+çeyrek+Bilançoya+Göre+Hisse+Ederi+Hesaplama.xlsx",
    "/home/user/BorsaPortal/Temel/2.+çeyrek+Bilançoya+Göre+Hisse+Ederi+Hesaplama.xlsx",
    "/home/user/BorsaPortal/Temel/3.+çeyrek+Bilançoya+Göre+Hisse+Ederi+Hesaplama.xlsx",
    "/home/user/BorsaPortal/Temel/4.+çeyrek+Bilançoya+Göre+Hisse+Ederi+Hesaplama.xlsx",
    "/home/user/BorsaPortal/Temel/YIL+SONU+TAHMİNİ+KÂRA+GÖRE+HİSSE+EDERİ+HESAPLAMA.xlsx",
    "/home/user/BorsaPortal/Temel/2023-UZUN+VADE+HEDEF+HESAPLAMA+TABLOSU.xlsx"
]

for excel_file in excel_files:
    if os.path.exists(excel_file):
        analyze_excel_file(excel_file)
    else:
        print(f"File not found: {excel_file}")
